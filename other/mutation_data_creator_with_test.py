"""
Module Name: FastaDataProcessor
Author: [Your Name]
Contact: [Your Email]

Description:
    This module processes DNA segments and associated metadata from a .fasta file and metadata file. It interacts with
    a database to create or associate isolates, search for references, and store mutation information.
    The module writes data to the Segment, SegmentMutations, Mutation, Isolate, and Location tables.

    If the tables do not exist in the database, they are created at the beginning of the process.
"""
import datetime
# Import necessary libraries
import gc
import os
import shutil
import re

import openpyxl
import pandas
import sqlite3
import variant_calling

codon_table = {
    "UUU": "F", "UUC": "F", "UUA": "L", "UUG": "L",
    "UCU": "S", "UCC": "S", "UCA": "S", "UCG": "S",
    "UAU": "Y", "UAC": "Y", "UAA": "*", "UAG": "*",
    "UGU": "C", "UGC": "C", "UGA": "*", "UGG": "W",
    "CUU": "L", "CUC": "L", "CUA": "L", "CUG": "L",
    "CCU": "P", "CCC": "P", "CCA": "P", "CCG": "P",
    "CAU": "H", "CAC": "H", "CAA": "Q", "CAG": "Q",
    "CGU": "R", "CGC": "R", "CGA": "R", "CGG": "R",
    "AUU": "I", "AUC": "I", "AUA": "I", "AUG": "M",
    "ACU": "T", "ACC": "T", "ACA": "T", "ACG": "T",
    "AAU": "N", "AAC": "N", "AAA": "K", "AAG": "K",
    "AGU": "S", "AGC": "S", "AGA": "R", "AGG": "R",
    "GUU": "V", "GUC": "V", "GUA": "V", "GUG": "V",
    "GCU": "A", "GCC": "A", "GCA": "A", "GCG": "A",
    "GAU": "D", "GAC": "D", "GAA": "E", "GAG": "E",
    "GGU": "G", "GGC": "G", "GGA": "G", "GGG": "G",
}

# Load the workbook and select the sheet
file_path = "/Database/data_scripts/example_data/Sequences_Example_Italy_H5N1.xlsm"
workbook = openpyxl.load_workbook(file_path, data_only=True)


def process_testing(aligned_and_translated_cds, sheet_name, col_name, i, row_name):
    sheet = workbook[sheet_name]

    # Initialize row and column indices
    row_index = None
    col_index = None

    # Find the row index for the row_name
    for row in sheet.iter_rows(1, sheet.max_row):
        if row[0].value == row_name:
            row_index = row[0].row
            break

    # Find the column index for the col_name
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value.startswith(col_name+":") and str(i) == extract_last_number(col[0].value):
            col_index = col[0].column
            break

    # Retrieve the value at the intersection of the found row and column
    if row_index is not None and col_index is not None:
        with open("test_results", 'a+') as file:
            if sheet.cell(row=row_index, column=col_index).value != aligned_and_translated_cds[i-1]:
                test_str = f"{row_name}_{col_name}: PROBLEM"
                file.write(test_str + '\n')
                return -1
    else:
        return None


def extract_last_number(string):
    # Regular expression pattern to find the number between two letters
    pattern = r'[A-Za-z]([0-9]+)[A-Za-z]'

    # Search for the pattern in the string
    match = re.search(pattern, string)
    if match:
        # If pattern is found, extract the number
        return match.group(1)
    else:
        return None  # Return None if no match is found


test_dict = {
    "A/black-headed_gull/Italy/23VIR1099-4/2023|EPI_ISL_16979821||A_/_H5N1": "A/black-headed_gull/Italy/23VIR1099-4/2023_H5N1_2023-02-03",
    "A/buzzard/Italy/21VIR11899-5/2021|EPI_ISL_8882216||A_/_H5N1": "A/buzzard/Italy/21VIR11899-5/2021_H5N1_2021-12-27",
    "A/chicken/Italy/21VIR10264/2021|EPI_ISL_14760645||A_/_H5N1": "A/chicken/Italy/21VIR10264/2021_H5N1_2021-11-28",
    "A/chicken/Italy/21VIR10814-1/2021|EPI_ISL_14760691||A_/_H5N1": "A/chicken/Italy/21VIR10814-1/2021_H5N1_2021-12-07",
    "A/guinea_fowl/Italy/21VIR10351/2021|EPI_ISL_14760638||A_/_H5N1": "A/guinea_fowl/Italy/21VIR10351/2021_H5N1_2021-11-26",
    "A/mallard/Italy/22VIR9762-1/2022|EPI_ISL_18652553||A_/_H5N1": "A/mallard/Italy/22VIR9762-1/2022_H5N1_2022-10-12",
    "A/turkey/Italy/21VIR10851/2021|EPI_ISL_8882172||A_/_H5N1": "A/turkey/Italy/21VIR10851/2021_H5N1_2021-12-06",
    "A/turkey/Italy/21VIR9210-1/2021|EPI_ISL_7733647||A_/_H5N1": "A/turkey/Italy/21VIR9210-1/2021_H5N1_2021-11-05",
    "A/turkey/Italy/23VIR1887-10/2023|EPI_ISL_17679748||A_/_H5N1": "A/turkey/Italy/23VIR1887-10/2023_H5N1_2023-03-06",
    "A/turkey/Italy/24VIR1352-4/2024|EPI_ISL_18956169||A_/_H5N1": "A/turkey/Italy/24VIR1352-4/2024_H5N1_2024-02-20"
}

test_dict_2 = {
    "PB2": "PB2",
    "PB1": "PB1",
    "PB1-F2": "PB1-F2",
    "PA": "PA",
    "HA1": "HA1-5",
    "HA2": "HA2-5",
    "NP": "NP",
    "NA": "NA",
    "M1": "M1",
    "NS": "NS-1"
}


# Define a class for FastaDataProcessor
class MutationDatabaseHandler:

    conn = sqlite3.connect('../Database/data_scripts/database/thesis.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    def __init__(self):

        gc.enable()  # Enable garbage collection

        try:
            shutil.rmtree("../Database/data_scripts/tmp")
        except FileNotFoundError:
            pass

        try:
            os.mkdir("../Database/data_scripts/tmp")
            os.mkdir("../Database/data_scripts/tmp/dna")
            os.mkdir("../Database/data_scripts/tmp/protein")
        except FileExistsError:
            pass

        self.fasta_file = None
        self.metadata_file = None
        self.metadata_dict = {}
        self.dna_fasta_by_ref = {}
        self.dna_fasta_dict = {}
        self.protein_fasta_by_annotation = {}  # Only one not cleared at runtime
        self.protein_fasta_dict = {}

        self.log_file = open("../Database/data_scripts/tmp/bad_sequences.log", "a")

        self.create_tables_if_not_exist()

    def __del__(self):
        self.log_file.close()

    def execute_data_flow(self, folder):

        self.read_fasta_and_metadata(folder)
        self.get_metadata_dict(self.metadata_file)
        self.organize_dna_fasta_by_reference()

        # Free up RAM
        del self.metadata_dict
        self.fasta_file.close()
        del self.metadata_file
        gc.collect()

        for serotype, segment in [(key.split('_')[0], key.split('_')[1]) for key in self.dna_fasta_by_ref.keys()]:
            dna_reference = self.get_reference(serotype, segment)

            if dna_reference:

                annotations = self.get_annotations(segment)
                dna_insertions_file = self.align_dna_sequences(serotype, segment, dna_reference["dna_fasta"])

                for annotation in annotations:

                    self.organize_protein_fasta_by_annotation(serotype, segment, annotation,
                                                              self.dna_fasta_dict, dna_insertions_file)

                    if not (f"{serotype}_{segment}_{annotation["annotation_id"]}"
                            in self.protein_fasta_by_annotation.keys()):
                        continue

                    reference_protein_fasta = self.get_protein_reference(annotation, dna_reference["dna_fasta"])
                    protein_insertion_file = (
                        self.align_protein_sequences(serotype, segment, annotation, reference_protein_fasta))

                    for header, aligned_protein_fasta in self.protein_fasta_dict.items():

                        header_dict = self.get_header_dict(header)

                        # Just a double check to be 100% sure
                        if not self.does_segment_exists(header_dict):
                            seg_id = self.create_segment(header_dict)  # Here so that quitting midway is not a problem
                        else:
                            seg_id = self.get_segment_id(header_dict)

                        mutations = self.get_mutations(
                            header, aligned_protein_fasta, reference_protein_fasta, protein_insertion_file)

                        for mutation in mutations:
                            mut_id = self.create_mutation(header_dict['segment_serotype'], annotation, mutation)
                            self.create_segment_mutation(seg_id, dna_reference["reference_id"], mut_id)

                    #self.conn.commit()  # Save mutations and segments found

            # Free up RAM
            del self.dna_fasta_by_ref[f"{serotype}_{segment}"]
            gc.collect()

        self.conn.close()

    """ --- EXECUTION FUNCTIONS --- """

    def organize_dna_fasta_by_reference(self):

        header, sequence = None, ""
        for line in self.fasta_file.readlines():
            if line.startswith('>'):

                if header:
                    header_dict = self.get_header_dict(header)
                    if not self.does_segment_exists(header_dict):
                        if self.is_valid_sequence(sequence):
                            if self.create_isolate(header_dict["isolate_id"]) != -1:
                                self.add_to_dna_fasta_by_ref_dict(header, sequence)

                header = line.strip()
                sequence = ""
            else:
                sequence += line.strip()

        if header:
            header_dict = self.get_header_dict(header)
            if not self.does_segment_exists(header_dict):
                if self.is_valid_sequence(sequence):
                    if self.create_isolate(header_dict["isolate_id"]) != -1:
                        self.add_to_dna_fasta_by_ref_dict(header, sequence)

        #self.conn.commit()  # Save isolates

    def organize_protein_fasta_by_annotation(self, serotype, segment, annotation,
                                             aligned_dna_fastas, dna_insertions_file):

        try:
            insertions_file = open(dna_insertions_file, "r")
            insertions_csv = pandas.read_csv(insertions_file, on_bad_lines='skip', low_memory=False)
        except:
            insertions_file = None
            insertions_csv = None

        for aligned_header, aligned_dna_fasta in aligned_dna_fastas.items():

            # Remove sequences with too many "-"
            if self.is_bad_sequence(aligned_dna_fasta):
                log_entry = f"Bad sequence, too few nucleotides: {aligned_header}\n"
                self.log_file.write(log_entry)
                continue

            aligned_dna_fasta = aligned_dna_fasta.upper().replace("T", "U").replace("\n", "")

            if insertions_file:
                populated_dna_fasta, n_added_nucleotides = self.reinsert_insertions(aligned_header,
                                                                                    aligned_dna_fasta,
                                                                                    annotation, insertions_csv)
            else:
                populated_dna_fasta, n_added_nucleotides = aligned_dna_fasta, 0

            aligned_cds = self.get_aligned_cds(populated_dna_fasta, annotation, n_added_nucleotides)
            aligned_cds.replace("-", "")

            try:
                cds_aminoacid = self.dna2aminoacid(aligned_cds)
            except ValueError:
                log_entry = f"CDS not divisible by 3: {aligned_header}\n"
                self.log_file.write(log_entry)
                continue

            for i in range(len(cds_aminoacid)):
                try:
                    if process_testing(cds_aminoacid, "Mutations",
                                       test_dict_2[annotation["annotation_name"]], i,
                                       test_dict["|".join(aligned_header.split('|')[2:])]) == -1:
                        print("Problem")
                except KeyError:
                    pass

            annotation_id = annotation["annotation_id"]
            try:
                self.protein_fasta_by_annotation[f"{serotype}_{segment}_{annotation_id}"] \
                    += f"{aligned_header}\n{cds_aminoacid}\n"
            except KeyError:
                self.protein_fasta_by_annotation[f"{serotype}_{segment}_{annotation_id}"] = \
                    f"{aligned_header}\n{cds_aminoacid}\n"

    def align_dna_sequences(self, serotype, segment, reference_dna_fasta):

        dna_path = "../Database/data_scripts/tmp/dna/"
        dna_reference_file = self.create_fasta_file(dna_path,
                                                    f"{serotype}_{segment}_reference",
                                                    f">{serotype}_{segment}_reference\n{reference_dna_fasta}")

        dna_target_file = self.create_fasta_file(dna_path,
                                                 f"{serotype}_{segment}_target",
                                                 self.dna_fasta_by_ref[f"{serotype}_{segment}"])

        aligned_file_path = dna_path + f"{serotype}_{segment}_aligned.fasta"
        self.execute_augur(dna_reference_file, dna_target_file, aligned_file_path)
        self.populate_dna_fasta_dict(aligned_file_path)

        insertions_file = aligned_file_path.replace(".fasta", ".fasta.insertions.csv") \
            if (os.path.isfile(aligned_file_path.replace(".fasta", ".fasta.insertions.csv"))) else None
        return insertions_file

    def align_protein_sequences(self, serotype, segment, annotation, reference_protein_fasta):

        protein_path = "../Database/data_scripts/tmp/protein/"
        annotation_id = annotation["annotation_id"]
        protein_reference_file = self.create_fasta_file(
            protein_path,
            f"{serotype}_{segment}_{annotation_id}_reference",
            f">{serotype}_{segment}_{annotation_id}_reference\n{reference_protein_fasta}")

        protein_target_file = (
            self.create_fasta_file(protein_path,
                                   f"{serotype}_{segment}_{annotation_id}_target",
                                   self.protein_fasta_by_annotation[f"{serotype}_{segment}_{annotation_id}"]))

        protein_aligned_file = protein_path + f"{serotype}_{segment}_{annotation_id}_aligned.fasta"
        self.execute_augur(protein_reference_file, protein_target_file, protein_aligned_file)
        self.populate_protein_fasta_dict(protein_aligned_file)

        # Free up RAM
        del self.protein_fasta_by_annotation[f"{serotype}_{segment}_{annotation_id}"]
        gc.collect()

        insertions_file = protein_aligned_file.replace(".fasta", ".fasta.insertions.csv") \
            if (os.path.isfile(protein_aligned_file.replace(".fasta", ".fasta.insertions.csv"))) else None
        return insertions_file

    @staticmethod
    def get_mutations(aminoacid_header, aminoacid_aligned, reference_aminoacid_fasta, insertion_file_path):
        variant_caller = variant_calling.VariantCaller(reference_aminoacid_fasta,
                                                       insertions_file_path=insertion_file_path)
        alignment_changes = variant_caller.call_variants(name=aminoacid_header, aligned=aminoacid_aligned,
                                                         aggregate_close_changes=False, one_based=False)
        return alignment_changes.subs + alignment_changes.dels + alignment_changes.ins

    """ --- FILE HANDLING FUNCTIONS --- """

    def read_fasta_and_metadata(self, folder):
        fasta_file = [file for file in os.listdir(folder) if file.endswith('.fasta')][0]
        metadata_file = [file for file in os.listdir(folder) if file.endswith('.xls')][0]
        self.fasta_file = open(f"{folder}/{fasta_file}", "r")
        self.metadata_file = pandas.read_excel(f"{folder}/{metadata_file}")

    @staticmethod
    def create_fasta_file(path, file_name, fasta):
        os.makedirs(path, exist_ok=True)
        with open(f"{path}{file_name}.fasta", "w+") as fasta_file:
            fasta_file.write(fasta)
        return f"{path}{file_name}.fasta"

    """ --- DATABASE FUNCTIONS --- """

    def get_reference(self, serotype_name, segment_type):
        serotype_id = self.get_serotype_id(serotype_name)
        if serotype_id is not None:
            try:
                self.cursor.execute("SELECT * FROM ReferenceSegment WHERE serotype_id = ? AND segment_type = ?",
                                    (serotype_id, segment_type))
                reference = self.cursor.fetchone()
                if reference: return reference
                else: return None
            except sqlite3.OperationalError: return None
        else: return None

    def get_serotype_id(self, serotype_name):
        try:
            self.cursor.execute("SELECT serotype_id FROM Serotype WHERE name = ?", (serotype_name,))
            serotype = self.cursor.fetchone()
            if serotype is not None: return serotype["serotype_id"]
            else: return None
        except sqlite3.OperationalError: return None

    def get_annotations(self, segment_type):
        try:
            self.cursor.execute("SELECT * FROM Annotation WHERE reference_id = "
                                "(SELECT reference_id FROM ReferenceSegment WHERE segment_type = ?)", (segment_type,))
            return self.cursor.fetchall()
        except sqlite3.OperationalError:
            return []

    def get_isolate_location_id(self, location_metadata):

        location_metadata = location_metadata.split("/")

        try:
            region = location_metadata[0].strip()
        except IndexError:
            region = "None"

        try:
            state = location_metadata[1].strip()
        except IndexError:
            state = "None"

        try:
            city = location_metadata[2].strip()
        except IndexError:
            city = "None"

        try:
            self.cursor.execute("SELECT location_id FROM Location "
                                "WHERE region = ? AND state = ? AND city = ?", (region, state, city))
            location = self.cursor.fetchone()
            if location:
                return location["location_id"]
        except sqlite3.OperationalError:
            pass

        # Insert Location into DB
        command = f"""INSERT INTO Location VALUES (?, ?, ?, ?)"""
        self.cursor.execute(command, (None, region, state, city))
        return self.cursor.lastrowid

    def get_segment_id(self, header_dict):
        try:
            self.cursor.execute("SELECT * FROM Segment WHERE epi_virus_name = ?", (f"{header_dict['segment_id']}",))
            segment = self.cursor.fetchone()
            if segment: return segment["segment_id"]
        except sqlite3.OperationalError:
            return False
        return False

    def create_isolate(self, isolate_epi):

        if not self.does_isolate_exists(isolate_epi):

            try:
                isolate_metadata = self.metadata_dict[isolate_epi]
            except KeyError:
                return -1

            host = isolate_metadata[17]
            collection_date = isolate_metadata[25]
            location_id = self.get_isolate_location_id(isolate_metadata[16])
            serotype_id = self.get_serotype_id(isolate_metadata[12].split("/")[-1].strip())

            # Insert Isolate into DB
            command = f"INSERT INTO Isolate VALUES (?, ?, ?, ?, ?, ?)"
            self.cursor.execute(command, (None, isolate_epi, serotype_id, host, collection_date, location_id))
            return 0
        else:
            return 0

    def create_segment(self, header_dict):

        isolate_id = header_dict["isolate_id"]
        segment_type = header_dict["segment_type"]
        segment_epi = header_dict["segment_epi"]
        virus_name = header_dict["virus_name"]
        segment_id = header_dict["segment_id"]

        command = f"""INSERT INTO Segment VALUES (?, ?, ?, ?, ?, ?)"""
        self.cursor.execute(command, (None, isolate_id, segment_type, segment_epi, virus_name, segment_id))

        # Return id for SegmentMutations
        return self.cursor.lastrowid

    def create_mutation(self, serotype, annotation, mutation):

        position = int(mutation[:mutation.find('_')]) + 1  # Comply with official position nomenclature
        ref = mutation[mutation.find('|') - 1:mutation.find('|')]
        alt = mutation[mutation.find('|') + 1:]
        mutation_name = f"{serotype}:{annotation["annotation_name"]}:{ref}{position}{alt}"

        mutation_id = self.does_mutation_exists(mutation_name)
        if mutation_id is None:

            command = f"""INSERT INTO Mutation VALUES (?, ?, ?, ?, ?, ?, ?, ?)"""
            self.cursor.execute(command,
                                (None, mutation_name, serotype, annotation["annotation_id"],
                                 annotation["annotation_name"], position, ref, alt))
            return self.cursor.lastrowid

        else:
            return mutation_id

    def create_segment_mutation(self, segment_id, reference_id, mutation_id):
        if not self.does_segment_mutation_exists(segment_id, reference_id, mutation_id):
            command = f"""INSERT INTO SegmentMutations VALUES (?, ?, ?)"""
            self.cursor.execute(command, (segment_id, reference_id, mutation_id))

    def create_tables_if_not_exist(self):

        # Location table
        create_location_table = """ 
        CREATE TABLE IF NOT EXISTS Location (
            location_id INTEGER PRIMARY KEY, region TEXT, state TEXT, city TEXT);"""

        # Isolate table
        create_isolate_table = """ 
        CREATE TABLE IF NOT EXISTS Isolate (
            isolate_id INTEGER PRIMARY KEY, isolate_epi TEXT, serotype_id INTEGER,
            host TEXT, collection_date DATE, location_id INTEGER,
            FOREIGN KEY (serotype_id) REFERENCES Serotype(serotype_id),
            FOREIGN KEY (location_id) REFERENCES Location(location_id));"""

        # Segment table
        create_segment_table = """ 
        CREATE TABLE IF NOT EXISTS Segment (
            segment_id INTEGER PRIMARY KEY, isolate_id INTEGER,
            segment_type TEXT, segment_epi TEXT, virus_name TEXT, epi_virus_name TEXT UNIQUE,
            FOREIGN KEY (isolate_id) REFERENCES Isolate(isolate_id));"""

        # Mutation table
        create_mutation_table = """
        CREATE TABLE IF NOT EXISTS Mutation (
            mutation_id INTEGER PRIMARY KEY, mutation_name TEXT UNIQUE,
            serotype_name TEXT, annotation_id, annotation_name TEXT, position INTEGER,
            ref TEXT, alt TEXT,
            FOREIGN KEY (annotation_id) REFERENCES Annotation(annotation_id));"""

        # SegmentMutations table
        create_segment_mutations_table = """
        CREATE TABLE IF NOT EXISTS SegmentMutations (
            segment_id INTEGER, reference_id INTEGER, mutation_id INTEGER,
            FOREIGN KEY (segment_id) REFERENCES Segment(segment_id),
            FOREIGN KEY (mutation_id) REFERENCES Mutation(mutation_id),
            FOREIGN KEY (reference_id) REFERENCES ReferenceSegment(reference_id),
            PRIMARY KEY (segment_id, mutation_id, reference_id));"""

        # Execute all create table statements
        create_table_statements = [
            create_location_table,
            create_isolate_table,
            create_segment_table,
            create_mutation_table,
            create_segment_mutations_table
        ]

        for create_statement in create_table_statements:
            self.cursor.execute(create_statement)

    def does_segment_exists(self, header_dict):
        try:
            self.cursor.execute("SELECT * FROM Segment WHERE epi_virus_name = ?", (f"{header_dict['segment_id']}",))
            if self.cursor.fetchone(): return True
        except sqlite3.OperationalError: return False
        return False

    def does_isolate_exists(self, isolate_epi):
        try:
            self.cursor.execute("SELECT * FROM Isolate WHERE isolate_epi = ?", (isolate_epi,))
            if self.cursor.fetchone(): return True
        except sqlite3.OperationalError: return False
        return False

    def does_mutation_exists(self, mutation_name):
        try:
            self.cursor.execute("SELECT * FROM Mutation WHERE mutation_name = ?", (mutation_name,))
            mut = self.cursor.fetchone()
            if mut is not None: return mut["mutation_id"]
        except sqlite3.OperationalError: return None
        return None

    def does_segment_mutation_exists(self, segment_id, reference_id, mutation_id):
        try:
            self.cursor.execute("SELECT * FROM SegmentMutations WHERE "
                                "segment_id = ? AND reference_id = ? AND mutation_id = ?",
                                (segment_id, reference_id, mutation_id))
            segm_mutations = self.cursor.fetchone()
            if segm_mutations is not None: return True
        except sqlite3.OperationalError: return False
        return False

    """ --- HELPER FUNCTIONS --- """

    def add_to_dna_fasta_by_ref_dict(self, header, sequence):
        segment = header.split("|")[1]
        serotype = header.split("|")[5][-4:]
        key = f"{serotype}_{segment}"
        entry = f"{header}\n{sequence}\n"

        if key in self.dna_fasta_by_ref:
            self.dna_fasta_by_ref[key] += entry
        else:
            self.dna_fasta_by_ref[key] = entry

    @staticmethod
    def execute_augur(reference_path, target_path, aligned_path):
        os.system("augur align"
                  f" --reference-sequence {reference_path}"
                  f" --sequences {target_path}"
                  f" --output {aligned_path}  --nthreads 4 --remove-reference")

    def populate_dna_fasta_dict(self, aligned_path):

        self.dna_fasta_dict.clear()
        header = None
        sequence = ""
        for line in open(aligned_path, "r").readlines():
            line = line.strip()
            if line.startswith('>'):
                if header is not None:
                    self.dna_fasta_dict[header] = sequence
                header = line
                sequence = ""
            else:
                sequence += line

        self.dna_fasta_dict[header] = sequence

    def populate_protein_fasta_dict(self, protein_aligned_file):

        self.protein_fasta_dict.clear()
        header = None
        sequence = ""
        for line in open(protein_aligned_file, "r").readlines():
            line = line.strip()
            if line.startswith('>'):
                if header is not None:
                    self.protein_fasta_dict[header] = sequence
                header = line
                sequence = ""
            else:
                sequence += line

        self.protein_fasta_dict[header] = sequence

    def get_metadata_dict(self, metadata):
        for row in metadata.values:
            self.metadata_dict[tuple(row)[0]] = tuple(row)

    @staticmethod
    def get_aligned_cds(aligned_fasta, annotation, n_added_nucleotides):
        aligned_cds = aligned_fasta.strip()[annotation["start_pos"] - 1: annotation["end_pos"] + n_added_nucleotides]
        return aligned_cds

    def get_protein_reference(self, annotation, reference_dna_fasta):
        reference_dna_fasta = reference_dna_fasta.upper().replace("T", "U").replace("\n", "")
        reference_cds = reference_dna_fasta.strip()[annotation["start_pos"] - 1: annotation["end_pos"]]
        reference_protein_fasta = self.dna2aminoacid(reference_cds)
        return reference_protein_fasta

    @staticmethod
    def get_header_dict(header):

        # Split the header line by '|'
        header_parts = header.split('|')

        # Extract the individual components
        segment_epi = header_parts[0][1:]
        segment_type = header_parts[1]
        virus_name = header_parts[2]
        isolate_id = header_parts[3]
        # If there's an additional part for the serotype, extract it
        segment_serotype = header_parts[-1].split('_')[-1] if '_' in header_parts[-1] else None

        # Construct the segment_id
        segment_id = f"{segment_epi}:{virus_name}"

        # Create the dictionary
        return {
            "segment_type": segment_type,
            "segment_epi": segment_epi,
            "virus_name": virus_name,
            "segment_id": segment_id,
            "isolate_id": isolate_id,
            "segment_serotype": segment_serotype
        }

    @staticmethod
    def dna2aminoacid(sequence):

        # Ensure the sequence length is divisible by 3
        if len(sequence) % 3 != 0:
            raise ValueError("Sequence length must be a multiple of 3")

        protein_sequence = ""
        for i in range(0, len(sequence), 3):
            codon = sequence[i:i+3]
            if codon in codon_table:
                amino_acid = codon_table[codon]
                protein_sequence += amino_acid
            else:
                raise ValueError("Invalid codon: " + codon)

        return protein_sequence

    @staticmethod
    def is_valid_sequence(s):
        valid = 'actgu'
        for letter in s:
            if letter not in valid:
                return False
        return True

    @staticmethod
    def is_bad_sequence(s):
        hyphen_count = s.count('-')
        total_length = len(s)
        hyphen_percentage = (hyphen_count / total_length) * 100 if total_length > 0 else 0
        return hyphen_percentage > 60

    def reinsert_insertions(self, aligned_header, aligned_fasta, annotation, dna_insertions_file):

        n_added_nucleotides = 0
        row = dna_insertions_file.loc[dna_insertions_file['strain'] == aligned_header.replace(">", "")]
        if row.empty:
            return aligned_fasta, n_added_nucleotides

        row_dict = row.to_dict(orient='records')[0]
        row_dict = {key: value for i, (key, value) in enumerate(row_dict.items()) if i != 0}

        for insertion, nucleotides in row_dict.items():
            pos = int(insertion[insertion.index("pos ") + len("pos "):])
            n_nucleotides = int(insertion[insertion.index("insertion: ") + len("insertion: "): insertion.index("bp")])

            if pos == annotation["start_pos"]:
                pos, nucleotides, n_nucleotides = self.check_start_pos(pos, nucleotides, n_nucleotides)

            if pos == annotation["end_pos"]:
                pos, nucleotides, n_nucleotides = self.check_end_pos(pos, nucleotides)

            if annotation["start_pos"] <= pos <= annotation["end_pos"] and pandas.notna(nucleotides):
                aligned_fasta = (aligned_fasta[:pos - 1 + n_added_nucleotides] + nucleotides +
                                 aligned_fasta[pos - 1 + n_added_nucleotides:])
                n_added_nucleotides += n_nucleotides

        return aligned_fasta, n_added_nucleotides

    @staticmethod
    def check_start_pos(pos, nucleotides, n_nucleotides):
        start_codon = "AUG"
        start_codon_index = nucleotides.find(start_codon)
        if start_codon_index == -1:
            return 0, "", 0  # Ignore insertion, cause it doesn't have start codon
        return pos, nucleotides[start_codon_index:], n_nucleotides - start_codon_index

    @staticmethod
    def check_end_pos(pos, nucleotides):

        stop_codons = ["UAA", "UAG", "UGA"]
        min_stop_codon_index = -1
        for stop_codon in stop_codons:
            stop_codon_index = nucleotides.find(stop_codon)
            if stop_codon_index != -1:
                if min_stop_codon_index == -1 or stop_codon_index < min_stop_codon_index:
                    min_stop_codon_index = stop_codon_index

        if min_stop_codon_index == -1:
            return 0, "", 0  # Ignore insertion, because it doesn't have a stop codon

        return pos, nucleotides[:min_stop_codon_index + 3], min_stop_codon_index + 3


if __name__ == "__main__":
    DBHandler = MutationDatabaseHandler()
    DBHandler.execute_data_flow("/home/luca/Desktop/ThesisTODO/Database/data_scripts/segments_data/TEST")
